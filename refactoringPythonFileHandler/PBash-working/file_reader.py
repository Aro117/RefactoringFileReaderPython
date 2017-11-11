from abc import ABCMeta, abstractmethod
import sys
import csv
import re
import openpyxl
from validator import Validator


class FileReader(metaclass=ABCMeta):
    @abstractmethod
    def read_file(self, filename):
        pass


class CSVReader(FileReader):
    def __init__(self, new_validator):
        self.validator = new_validator

    def read_file(self, filename):
        try:
            with open(filename) as f_obj:
                reader = csv.DictReader(f_obj, delimiter=',')
                the_list = []
                for line in reader:
                    employee = dict()
                    employee["EMPID"] = line["emp_id"]
                    employee["GENDER"] = line["gender"]
                    employee["AGE"] = line["age"]
                    employee["SALES"] = line["sales"]
                    employee["BMI"] = line["bmi"]
                    employee["SALARY"] = line["salary"]
                    employee["BIRTHDAY"] = line["birthday"]

                    if self.validator.check_line(employee):
                        the_list.append(employee)
                    else:
                        print('Entry failed validation', file=sys.stderr)
                if self.validator.check_data_set(the_list):
                    return the_list
                else:
                    print("There were no valid entries in the file", file=sys.stderr)
                    return False
        except FileNotFoundError:
            print('The file was not found', file=sys.stderr)
            return False


class TXTReader(FileReader):
    def __init__(self, new_validator):
        self.validator = new_validator

    def read_file(self, filename):
        try:
            file = open(filename, "r")
        except FileNotFoundError:
            print('The file was not found', file=sys.stderr)
            return False
        the_list = []
        for line in file:
            dictionary = {}
            entries = line.split(";")
            for entry in entries:
                if len(entry.split("=")) == 2:
                    key = entry.split("=")[0]
                    value = entry.split("=")[1]
                    value = value.rstrip('\n')
                    dictionary[key] = value
                else:
                    print('The file was in an invalid format', file=sys.stderr)
                    return False
            if self.validator.check_line(dictionary):
                the_list.append(dictionary)
            else:
                print('Entry failed validation', file=sys.stderr)
        if self.validator.check_data_set(the_list):
            return the_list
        else:
            print("There were no valid entries in the file", file=sys.stderr)
            return False


class XLSXReader(FileReader):
    def __init__(self, new_validator):
        self.validator = new_validator

    def read_file(self, filename):
        try:
            wb = openpyxl.load_workbook(filename)
            sheet = wb.active
            the_list = []
            for x in range(2, 29):
                employee = dict()
                employee["EMPID"] = sheet.cell(column=1, row=x).value
                employee["GENDER"] = sheet.cell(column=2, row=x).value
                employee["AGE"] = sheet.cell(column=3, row=x).value
                employee["SALES"] = sheet.cell(column=4, row=x).value
                employee["BMI"] = sheet.cell(column=5, row=x).value
                employee["SALARY"] = sheet.cell(column=6, row=x).value
                employee["BIRTHDAY"] = sheet.cell(column=7, row=x).value
                if self.validator.check_line(employee):
                    the_list.append(employee)
                else:
                    print('Entry failed validation', file=sys.stderr)
            if self.validator.check_data_set(the_list):
                return the_list
            else:
                print("There were no valid entries in the file", file=sys.stderr)
                return False
        except FileNotFoundError:
            print("File not found!", file=sys.stderr)
            return False
